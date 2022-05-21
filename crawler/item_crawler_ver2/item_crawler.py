from selenium import webdriver
from selenium.webdriver.common.by import By
from easydict import EasyDict
from utils import *
import openpyxl
import os

import pandas as pd
from tqdm import tqdm

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument(argument='--headless') 
chrome_options.add_argument(argument='--no-sandbox')
chrome_options.add_argument(argument='--disable-dev-shm-usage')

URL_PATH = "https://www.musinsa.com/app/styles/lists"
driver = webdriver.Chrome('/opt/ml/workspace/crawler/chromedriver',chrome_options=chrome_options)
driver.get(URL_PATH)
driver.implicitly_wait(2) #페이지를 로딩하는 시간동안 대기

# 🚀 크롤링 완료된 정보를 저장할 excel sheet_codi 지정
wb_item = openpyxl.Workbook()
sheet_item = wb_item.active
sheet_item.append(["id",  "name", "big_class", "mid_class", "brand", "serial_number", "gender",
                   "season", "cum_sale", "view", "likes", "rating", "price", "url", "img_url", "codi_id"])

wb_item_color = openpyxl.Workbook()
sheet_item_color = wb_item_color.active
sheet_item_color.append(["id", "color"])

wb_item_size = openpyxl.Workbook()
sheet_item_size = wb_item_size.active
sheet_item_size.append(["id", "size"])

wb_item_tag = openpyxl.Workbook()
sheet_item_tag = wb_item_tag.active
sheet_item_tag.append(["id", "tag"])

wb_item_four_season = openpyxl.Workbook()
sheet_item_four_season = wb_item_four_season.active
sheet_item_four_season.append(["id", "four_season"])

wb_item_fit = openpyxl.Workbook()
sheet_item_fit = wb_item_fit.active
sheet_item_fit.append(["id", "fit"])

wb_item_buy_age = openpyxl.Workbook()
sheet_item_buy_age = wb_item_buy_age.active
sheet_item_buy_age.append(["id", "buy_age_18", "buy_age_19_23", "buy_age_24_28", 
                           "buy_age_29_33", "buy_age_34_39", "buy_age_40"])

wb_item_buy_gender = openpyxl.Workbook()
sheet_item_buy_gender = wb_item_buy_gender.active
sheet_item_buy_gender.append(["id", "buy_men", "buy_women"])


button = driver.find_element(By.CSS_SELECTOR, "button.global-filter__button--mensinsa")
button.click()

#container =  driver.find_elements(By.CSS_SELECTOR, "ul.style-list > li")
#BASE_CODI_URL = "https://www.musinsa.com/app/styles/views/"
#codi = []
#for codi_element in container:
#    img_src = codi_element.find_element(By.CSS_SELECTOR, 'div.style-list-item__thumbnail > a').get_attribute('onclick')
#    codi_id = img_src.split("'")[1]
#    codi_url = BASE_CODI_URL + codi_id
#    codi.append((codi_id, codi_url))

codi_info = pd.read_excel("/opt/ml/workspace/crawler/codi_crawler/asset/codi.xlsx")
codi_urls = codi_info["url"].to_list()
codi_ids = codi_info["id"].to_list()

#for codi_id, codi_url in codi:
#    print(f"Crawling for CODI ID : {codi_id}\n")
cnt = 0
for codi_id, codi_url in zip(codi_ids[235:], codi_urls[235:]) :
    print(f"Crawling for CODI URL : {codi_url}\n")
    print(f"{cnt} out of 600 codi crawled...")

    try :
        driver.get(codi_url)
    except :
        print("이 에러가 발생하면 다음 코디부터 따로 크롤링 해주시길 바랍니다!")
        continue
    
    driver.implicitly_wait(1)
    item_list = driver.find_elements(By.CSS_SELECTOR, 'div.styling_list > div.swiper-slide')
    item_urls = []
    
    for item in item_list:
        item_url = item.find_element(By.CSS_SELECTOR, "a.brand_item").get_attribute('href')
        item_urls.append(item_url)

    for item_url in item_urls:
        try : 
            driver.get(item_url)
        except :
            continue
        driver.implicitly_wait(0.5) #페이지를 로딩하는 시간동안 대기
        print(f"Crawling item : {item_url}")

        id = item_url.split('/')[-2]
        name = driver.find_element(By.CSS_SELECTOR, "span.product_title > em").text
        category = driver.find_elements(By.CSS_SELECTOR, "p.item_categories > a")
        big_class = get_big_class(category)
        mid_class = get_mid_class(category)
        product_info = driver.find_elements(By.CSS_SELECTOR, "ul.product_article > li > p.product_article_contents > strong")
        brand = get_brand(product_info)
        serial_number = get_serial_number(product_info)
        season = get_season(product_info)
        gender = get_gender(driver)
        view = get_view(driver)
        cum_sale = get_cum_sale(driver)
        likes = get_likes(driver)
        rating = get_rating(driver)  
        price = get_price(driver)
        img_url = driver.find_element(By.CSS_SELECTOR, "div.product-img > img").get_attribute('src')
        
        try: menu = driver.find_elements(By.CSS_SELECTOR, "div#goods_opt_area > select")
        except: menu = None
        color_list = get_color(menu)
        size_list = get_size(menu)
        tags_list = get_tags_list(driver)
        four_season_list, fit_list = get_fs_and_fit(driver)      
        buy_age_list = get_buy_age_list(driver)
        buy_gender_list = get_buy_gender_list(driver)
        
        sheet_item.append([id,  name, big_class, mid_class, brand, serial_number, gender,
                   season, cum_sale, view, likes, rating, price, item_url, img_url, codi_id])
        print([id,  name, big_class, mid_class, brand, serial_number, gender,
                   season, cum_sale, view, likes, rating, price, item_url, img_url, codi_id])
        
        print(f'color_list: {color_list}')
        if color_list:
            for color in color_list:
                sheet_item_color.append([id, color])
            
        print(f'size_list: {size_list}')
        if size_list:
            for size in size_list:
                sheet_item_size.append([id, size])
            
        print(f'tags_list: {tags_list}')
        if tags_list:
            for tag in tags_list:
                sheet_item_tag.append([id, tag])
            
        print(f'four_season_list: {four_season_list}')
        if four_season_list:
            for four_season in four_season_list:
                sheet_item_four_season.append([id, four_season])
            
        print(f'fit_list: {fit_list}')
        if fit_list:
            for fit in fit_list:
                sheet_item_fit.append([id, fit])
            
        print(f'buy_age_list: {buy_age_list}')
        if buy_age_list:
            sheet_item_buy_age.append([id]+buy_age_list)
        
        print(f'buy_gender_list: {buy_gender_list}')
        if buy_gender_list:
            sheet_item_buy_gender.append([id]+buy_gender_list)
        
        os.makedirs('./asset2', exist_ok=True)
        wb_item.save("./asset2/item.xlsx")
        wb_item_color.save("./asset2/item_color.xlsx")
        wb_item_size.save("./asset2/item_size.xlsx")
        wb_item_tag.save("./asset2/item_tag.xlsx")
        wb_item_four_season.save("./asset2/item_four_season.xlsx")
        wb_item_fit.save("./asset2/item_fit.xlsx")
        wb_item_buy_age.save("./asset2/item_buy_age.xlsx")
        wb_item_buy_gender.save("./asset2/item_buy_gender.xlsx")

        print()
    cnt += 1


driver.close()

